"""
==========================================================
  Vitalsy Elite AI  —  brain.py  v4.0
  Blood Report Analysis Backend  (FastAPI + Groq)

  pip install fastapi uvicorn groq firebase-admin
              python-docx fpdf2 pdfplumber openpyxl
              python-multipart python-dotenv

  RUN locally:  python brain.py
  RUN on Render: uvicorn brain:app --host 0.0.0.0 --port 10000
==========================================================
"""

import io, os, re, json, traceback, csv, base64
import pdfplumber
from groq import Groq
from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, HTMLResponse, FileResponse
from pydantic import BaseModel
from datetime import datetime, timezone
from fpdf import FPDF
from docx import Document

load_dotenv()

GROQ_API_KEY  = os.getenv("GROQ_API_KEY", "")
FIREBASE_CRED = os.getenv("FIREBASE_CRED", "serviceAccountKey.json")

if not GROQ_API_KEY:
    raise RuntimeError(
        "\n\n❌  GROQ_API_KEY not found!\n"
        "    Add it as an Environment Variable on Render.\n"
        "    Or create a .env file locally with: GROQ_API_KEY=your_key\n"
    )

groq_client       = Groq(api_key=GROQ_API_KEY)
GROQ_MODEL        = "llama-3.3-70b-versatile"
GROQ_VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"

# ─────────────────────────────────────────────
# FIX 1: SYSTEM PROMPT — stronger no-markdown rule
# ─────────────────────────────────────────────
SYSTEM_PROMPT = """
You are Vitalsy Elite AI — a sharp, warm, and experienced medical specialist built by the Vitalsy team.

=== WHO YOU ARE ===
- AI medical assistant specialising exclusively in blood test report analysis.
- Built by the Vitalsy team to make healthcare understandable for everyone — educated or not.
- Version 1. More report types (MRI, X-ray, urine, ECG) coming soon.
- If asked who you are, who built you, what you do, what is special about you — answer naturally and confidently, like a real person would.

=== FORMATTING — THIS IS CRITICAL, DO NOT IGNORE ===
ABSOLUTE RULE: Do NOT use any markdown formatting. Ever. Not even once.
- No ** bold **, no *italic*, no ~~strikethrough~~
- No ### headings, no ## headings, no # headings
- No | table | pipes | or table formatting
- No --- or === dividers
- No backticks or code blocks
Write in plain text only. Use a blank line to separate sections. Use a dash (-) for bullet points ONLY when listing items.
If you break this rule, the app will display broken code to the user. Plain text only.

=== RESPONSE STYLE ===
- Write like a real doctor texting a patient. Not an essay. Not a wall of paragraphs.
- Mix short sentences, dash bullet points, and brief paragraphs naturally.
- NEVER write more than 2 sentences in a row without a line break or a bullet point.
- SHORT responses for greetings, simple questions, clarifications. Max 5 lines.
- STRUCTURED responses only for actual blood report analysis.
- Split long responses into clearly labelled sections with a blank line between each.
- Do not repeat the same format every reply. Adapt to the conversation.

=== GREETING ===
When user says hi/hello/hey — ONE warm sentence, then 2-3 short lines about what they can do.
Max 5 lines total. Example:

Hey! Good to have you here. I am Vitalsy AI — your blood report specialist.
You can upload a blood report and I will break it down for you, or just describe what you are feeling and we can go from there.

=== WHEN ASKED WHAT YOU CAN DO ===
Short list, no markdown. Example:

Here is what I can help with:

- Read and explain your blood test report in plain language
- Flag values that are High, Low, or Normal
- Tell you what each result means for your body
- Give you diet and lifestyle tips based on your results
- Answer any questions about blood tests

Just upload your report or describe your symptoms and we will get started.

=== BLOOD REPORT ANALYSIS ===
When you receive blood report data, do a full analysis:

Start with 1-2 sentences on the overall picture. Reassure if mostly normal.

Then each parameter like this (plain text, no markdown):

Test Name
  Value: X  |  Normal Range: Y  |  Status: High / Low / Normal
  What it means: one plain English sentence
  Tip: one practical action if abnormal (skip if normal)

After all parameters, include these four sections (all mandatory):

OVERALL SUMMARY
3-4 plain sentences about the overall health picture.

WHAT TO DO NEXT
- bullet point 1
- bullet point 2
- bullet point 3

FOOD AND NUTRITION
Foods to eat more: (list with reasons based on specific values)
Foods to reduce: (list with reasons)
Simple meal idea — Breakfast: X | Lunch: Y | Dinner: Z

LIFESTYLE TIPS
- 3-4 specific habits based on results

End with: Please see a doctor before making any medical decisions based on this.

=== CLARIFICATION ===
If user describes symptoms without a report — acknowledge in 1 sentence, then ask age, gender, duration, conditions naturally.

=== SCOPE — BLOOD REPORTS ONLY ===
If user asks about non-blood reports (MRI, X-ray, urine, ECG, prescription, dental):
Reply: I am sorry, right now I can only analyse blood test reports. We are working hard to add more very soon — stay tuned! Sorry for the inconvenience and have a great day.

=== FOOD FOLLOW-UP ===
If user asks about food after a blood report was already analysed:
- Do NOT ask questions. Give specific foods immediately.
- Foods to eat more, foods to reduce, simple meal idea. Name actual foods.

If no blood report yet: give short general advice (5-6 lines), mention sharing their report for personalised advice.

=== RULES ===
- Never diagnose. Use: this may suggest, this could indicate, this is often seen in.
- If unsure — say so honestly and suggest seeing a doctor.
- Never be dismissive or cold.
- Keep follow-up answers connected to conversation context.
- Every reply should feel like it was written by a real person.
- NO MARKDOWN. PLAIN TEXT ONLY.
"""

# ─────────────────────────────────────────────
# FIREBASE
# ─────────────────────────────────────────────
db = None
try:
    import firebase_admin
    from firebase_admin import credentials, firestore
    if not firebase_admin._apps:
        RENDER_CRED = "/etc/secrets/serviceAccountKey.json"
        cred_path   = RENDER_CRED if os.path.exists(RENDER_CRED) else FIREBASE_CRED
        if os.path.exists(cred_path):
            cred = credentials.Certificate(cred_path)
            firebase_admin.initialize_app(cred)
            db = firestore.client()
            print(f"Firebase connected via: {cred_path}")
        else:
            print("Firebase cred file not found — chat history disabled.")
except Exception as e:
    print(f"Firebase not connected: {e}")

def save_chat(user_id, user_msg, ai_msg):
    if not db or user_id == "guest":
        return
    try:
        ref = db.collection("chats").document(user_id).collection("messages")
        ts  = datetime.now(timezone.utc)
        ref.add({"role": "user",      "content": user_msg, "timestamp": ts})
        ref.add({"role": "assistant", "content": ai_msg,   "timestamp": ts})
    except Exception as e:
        print("Firebase write error:", e)

# ─────────────────────────────────────────────
# FASTAPI
# ─────────────────────────────────────────────
app = FastAPI(title="Vitalsy Elite AI", version="4.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

# ─────────────────────────────────────────────
# MODELS
# ─────────────────────────────────────────────
class ChatRequest(BaseModel):
    message:  str
    user_id:  str  = "guest"
    history:  list = []

class AnalyzeRequest(BaseModel):
    text:    str
    user_id: str = "guest"

class PDFRequest(BaseModel):
    patient_name: str  = "Patient"
    conversation: list

SUPPORTED_EXTS = {".pdf",".docx",".doc",".txt",".csv",".xlsx",".xls",".rtf",".xml",".json",".png",".jpg",".jpeg"}
BLOCKED_EXTS   = {".mp3",".mp4",".wav",".avi",".mov",".mkv",".aac",".flac",".ogg",".wma",".wmv",".webm",".m4a",".m4v"}

# ─────────────────────────────────────────────
# AI HELPERS
# ─────────────────────────────────────────────
def normalise_history(history: list) -> list:
    messages = []
    for turn in history:
        role = turn.get("role", "user")
        if "content" in turn and turn["content"]:
            text = turn["content"]
        elif "parts" in turn:
            parts = turn["parts"]
            text = parts[0] if isinstance(parts, list) and parts else str(parts)
        else: continue
        if not text or not str(text).strip(): continue
        groq_role = "assistant" if role in ("model", "assistant") else "user"
        messages.append({"role": groq_role, "content": str(text).strip()})
    return messages

def call_groq(message: str, history: list, max_tokens: int = 2000) -> str:
    try:
        messages = [{"role": "system", "content": SYSTEM_PROMPT}]
        messages.extend(normalise_history(history))
        messages.append({"role": "user", "content": message})
        resp = groq_client.chat.completions.create(
            model=GROQ_MODEL, messages=messages, max_tokens=max_tokens, temperature=0.72
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        traceback.print_exc()
        return "I am having a little trouble connecting right now. Please try again in a moment."

# ─────────────────────────────────────────────
# FILE EXTRACTION
# ─────────────────────────────────────────────
def extract_text(file_bytes: bytes, filename: str) -> str:
    fname = filename.lower()
    if fname.endswith(".pdf"):
        parts = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t: parts.append(t)
        return "\n".join(parts)
    if fname.endswith((".docx", ".doc")):
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if fname.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    r = "\t".join(str(c) if c is not None else "" for c in row)
                    if r.strip(): rows.append(r)
            return "\n".join(rows)
        except: return ""
    if fname.endswith(".csv"):
        text = file_bytes.decode("utf-8", errors="ignore")
        try:
            reader = csv.reader(io.StringIO(text))
            return "\n".join(", ".join(row) for row in reader)
        except: return text
    if fname.endswith((".png", ".jpg", ".jpeg")):
        return "[IMAGE_FILE]"
    return file_bytes.decode("utf-8", errors="ignore")

# ─────────────────────────────────────────────
# FIX 2: VISION — no markdown, plain text only
# ─────────────────────────────────────────────
def analyse_image_with_groq(image_bytes: bytes, filename: str) -> str:
    ext       = os.path.splitext(filename)[1].lower().lstrip(".")
    mime_type = "image/png" if ext == "png" else "image/jpeg"
    b64       = base64.standard_b64encode(image_bytes).decode("utf-8")

    vision_instruction = """You are a medical blood report reading specialist.

CRITICAL FORMATTING RULE: Write in plain text only. No markdown whatsoever.
- No ** bold **, no ### headings, no | table | pipes, no --- dividers.
- Use plain text with line breaks and dash (-) for bullets only.
- If you use markdown, the app displays broken code to the user.

TASK: Read every value in this blood report image and give a full analysis.

Rules:
- Extract ALL test names and values. Even if slightly blurry, do your best.
- Do NOT say the image is unclear. Just read and analyse.
- For each parameter write exactly like this (plain text):

Test Name
  Value: X  |  Normal Range: Y  |  Status: High / Low / Normal
  What it means: one plain English sentence
  Tip: one action if abnormal (skip if normal)

After all parameters, write these four sections (all mandatory, plain text):

OVERALL SUMMARY
3-4 plain sentences. Reassure if mostly normal.

WHAT TO DO NEXT
- bullet 1
- bullet 2
- bullet 3

FOOD AND NUTRITION
Foods to eat more: list with reasons based on specific values
Foods to reduce: list with reasons
Simple meal idea — Breakfast: X | Lunch: Y | Dinner: Z

LIFESTYLE TIPS
- 3-4 specific habits based on the results

End with: Please consult a qualified doctor before making any medical decisions.

START the analysis immediately. Plain text only."""

    try:
        resp = groq_client.chat.completions.create(
            model=GROQ_VISION_MODEL,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}},
                    {"type": "text", "text": vision_instruction}
                ]
            }],
            max_tokens=2500,
            temperature=0.5,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        traceback.print_exc()
        return "I had trouble processing your image. Please try again, or type your blood values directly in the chat."

# ─────────────────────────────────────────────
# FIX 3: BUILD PDF — proper medical report, no chat transcript
# ─────────────────────────────────────────────
def clean_for_pdf(text: str) -> str:
    """Strip emojis, markdown, and non-latin chars for FPDF."""
    emoji_map = {
        "\U0001f534": "[HIGH]", "\U0001f7e1": "[LOW]", "\U0001f7e2": "[NORMAL]",
        "\u2705": "", "\u26a0\ufe0f": "[!]", "\u274c": "[X]",
    }
    for k, v in emoji_map.items():
        text = text.replace(k, v)
    text = re.sub(r'\*{1,3}(.*?)\*{1,3}', r'\1', text)
    text = re.sub(r'#{1,6}\s*', '', text)
    text = re.sub(r'_{1,2}(.*?)_{1,2}', r'\1', text)
    text = re.sub(r'\|', ' ', text)
    text = text.encode("latin-1", "replace").decode("latin-1")
    return text.strip()

def build_pdf(patient_name: str, conversation: list) -> bytes:
    """
    Build a proper medical report PDF.
    Extracts only AI analysis content — no chat conversation dump.
    """
    now = datetime.now()

    # Pull only AI messages (the actual analysis content)
    ai_messages = [
        m.get("content", "")
        for m in conversation
        if m.get("role") in ("assistant", "model")
        and len(m.get("content", "")) > 100  # skip short greetings
    ]
    analysis_content = "\n\n".join(ai_messages)

    # Ask Groq to generate a structured medical report from the analysis
    report_prompt = (
        f"You are writing a formal medical report document for a patient named {patient_name}.\n\n"
        f"Based on this blood report analysis consultation, write a clean structured medical report with these exact sections:\n\n"
        f"PATIENT INFORMATION\n"
        f"Patient Name: {patient_name}\n"
        f"Report Date: {now.strftime('%d %B %Y')}\n"
        f"Generated By: Vitalsy Elite AI\n\n"
        f"CLINICAL SUMMARY\n"
        f"Write 3-4 sentences summarising the patient overall health status based on the analysis.\n\n"
        f"PARAMETER ANALYSIS\n"
        f"For each blood parameter found in the analysis, write:\n"
        f"Parameter Name | Value | Normal Range | Status (Normal/High/Low) | Clinical Meaning\n\n"
        f"KEY FINDINGS\n"
        f"List the most important findings from the report. Note any abnormal values and their significance.\n\n"
        f"NUTRITION RECOMMENDATIONS\n"
        f"Based on the specific blood values, list foods to eat more and foods to avoid with reasons.\n\n"
        f"LIFESTYLE RECOMMENDATIONS\n"
        f"3-4 specific lifestyle suggestions based on the blood results.\n\n"
        f"MEDICAL DISCLAIMER\n"
        f"This report is AI-generated for informational purposes only. Always consult a licensed physician.\n\n"
        f"---\n"
        f"Consultation content to base this report on:\n{analysis_content[:4000]}\n\n"
        f"Rules: Plain text only. No markdown. No ** or ## or | tables. Write professionally."
    )

    try:
        structured = call_groq(report_prompt, [], max_tokens=2500)
    except Exception:
        structured = analysis_content

    structured = clean_for_pdf(structured)

    # ── COLOURS ──
    DARK_BLUE  = (0,  51, 102)
    MID_BLUE   = (0,  82, 142)
    LIGHT_BLUE = (230, 242, 255)
    WHITE      = (255, 255, 255)
    DARK_GRAY  = (50,  50,  50)
    MID_GRAY   = (100, 100, 100)
    LINE_GRAY  = (210, 210, 210)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=30)
    pdf.add_page()

    # ── HEADER ──
    pdf.set_fill_color(*DARK_BLUE)
    pdf.rect(0, 0, 210, 42, 'F')
    pdf.set_xy(10, 8)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Arial", "B", 22)
    pdf.cell(130, 12, "Vitalsy Elite AI", ln=False)
    pdf.set_font("Arial", "I", 9)
    pdf.set_xy(10, 22)
    pdf.cell(130, 6, "Blood Report Analysis | Medical Report", ln=False)
    pdf.set_xy(140, 9)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(60, 7, "MEDICAL REPORT", align="R", ln=True)
    pdf.set_xy(140, 17)
    pdf.set_font("Arial", size=8)
    pdf.cell(60, 5, f"Date: {now.strftime('%d %B %Y')}", align="R", ln=True)
    pdf.set_xy(140, 23)
    pdf.cell(60, 5, f"Report ID: VEA-{now.strftime('%Y%m%d%H%M%S')}", align="R", ln=True)

    # ── PATIENT BAR ──
    pdf.set_fill_color(*LIGHT_BLUE)
    pdf.rect(0, 42, 210, 16, 'F')
    pdf.set_text_color(*DARK_BLUE)
    pdf.set_font("Arial", "B", 10)
    pdf.set_xy(10, 46)
    pdf.cell(100, 6, f"Patient:  {patient_name}", ln=False)
    pdf.set_font("Arial", size=8)
    pdf.set_text_color(*MID_GRAY)
    pdf.set_x(110)
    pdf.cell(90, 6, "Powered by Vitalsy Elite AI  |  vitalsy.ai", align="R", ln=True)
    pdf.ln(7)

    pdf.set_draw_color(*MID_BLUE)
    pdf.set_line_width(0.6)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    # ── REPORT BODY ──
    SECTION_KEYWORDS = [
        "patient information", "clinical summary", "parameter analysis",
        "key findings", "nutrition recommendations", "lifestyle recommendations",
        "medical disclaimer", "overall summary", "what to do next",
        "food and nutrition", "lifestyle tips"
    ]

    for line in structured.split("\n"):
        line = line.strip()
        if not line:
            pdf.ln(2)
            continue

        ll         = line.lower()
        is_section = any(kw in ll for kw in SECTION_KEYWORDS) and len(line) < 80
        is_status  = any(t in line for t in ("[HIGH]", "[LOW]", "[NORMAL]", "High", "Low", "Normal"))
        is_bullet  = line.startswith("-") or line.startswith("*")

        if is_section:
            pdf.ln(3)
            pdf.set_fill_color(*MID_BLUE)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 7, f"   {line.upper()}", ln=True, fill=True)
            pdf.ln(2)
            pdf.set_text_color(*DARK_GRAY)

        elif "[HIGH]" in line or "| High |" in line or "Status: High" in line:
            pdf.set_fill_color(255, 230, 230)
            pdf.set_font("Arial", "B", 9)
            pdf.set_text_color(*DARK_GRAY)
            pdf.multi_cell(0, 6, line, fill=True)

        elif "[LOW]" in line or "| Low |" in line or "Status: Low" in line:
            pdf.set_fill_color(255, 250, 215)
            pdf.set_font("Arial", "B", 9)
            pdf.set_text_color(*DARK_GRAY)
            pdf.multi_cell(0, 6, line, fill=True)

        elif "[NORMAL]" in line or "| Normal |" in line or "Status: Normal" in line:
            pdf.set_fill_color(225, 255, 225)
            pdf.set_font("Arial", size=9)
            pdf.set_text_color(*DARK_GRAY)
            pdf.multi_cell(0, 6, line, fill=True)

        elif is_bullet:
            pdf.set_font("Arial", size=9)
            pdf.set_text_color(*DARK_GRAY)
            pdf.set_x(14)
            pdf.multi_cell(0, 5.5, "  " + line.lstrip("-* ").strip())

        else:
            pdf.set_font("Arial", size=9)
            pdf.set_text_color(*DARK_GRAY)
            pdf.multi_cell(0, 5.5, line)

    # ── FOOTER ──
    pdf.set_y(-28)
    pdf.set_fill_color(*DARK_BLUE)
    pdf.rect(0, pdf.get_y(), 210, 30, 'F')
    pdf.set_text_color(*WHITE)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 6, "Vitalsy Elite AI  |  Blood Report Analysis  |  vitalsy.ai", ln=True, align="C")
    pdf.set_font("Arial", "I", 7)
    pdf.cell(0, 5, "Disclaimer: This report is AI-generated for informational purposes only and does not constitute medical advice.", ln=True, align="C")
    pdf.cell(0, 4, "Please consult a licensed healthcare professional before making any medical decisions.", ln=True, align="C")

    result = pdf.output()
    if isinstance(result, bytearray):
        return bytes(result)
    return result

# ─────────────────────────────────────────────
# HTML FILE SERVING
# ─────────────────────────────────────────────
def get_file_path(filename: str):
    search_paths = [
        os.path.join(os.getcwd(), "templates", filename),
        os.path.join(os.getcwd(), filename),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), filename),
        os.path.join("/opt/render/project/src", "templates", filename),
        os.path.join("/opt/render/project/src", filename),
    ]
    for path in search_paths:
        if os.path.exists(path):
            return path
    return None

@app.get("/", response_class=FileResponse)
async def root():
    path = get_file_path("index.html")
    if not path: raise HTTPException(404, "index.html not found")
    return FileResponse(path)

@app.get("/login")
async def login_page():
    path = get_file_path("login.html")
    if not path: raise HTTPException(404, "login.html not found")
    return FileResponse(path)

@app.get("/signup")
async def signup_page():
    path = get_file_path("signup.html")
    if not path: raise HTTPException(404, "signup.html not found")
    return FileResponse(path)

@app.get("/main")
async def main_page():
    path = get_file_path("main.html")
    if not path: raise HTTPException(404, "main.html not found")
    return FileResponse(path)

@app.get("/{filename}.html")
async def serve_html(filename: str):
    path = get_file_path(f"{filename}.html")
    if not path: raise HTTPException(404, f"{filename}.html not found")
    return FileResponse(path)

# ─────────────────────────────────────────────
# API ROUTES — unchanged from your version
# ─────────────────────────────────────────────
@app.get("/api/status")
def api_status():
    return {"status": "Vitalsy Elite AI is running", "version": "4.0"}

@app.get("/health")
def health():
    return {"ok": True}

@app.options("/{rest_of_path:path}")
async def preflight(rest_of_path: str):
    return JSONResponse(content={}, status_code=200)

@app.post("/chat")
async def chat(request: ChatRequest):
    reply = call_groq(request.message, request.history)
    save_chat(request.user_id, request.message, reply)
    return {
        "reply": reply,
        "new_history_entry": [
            {"role": "user",      "content": request.message},
            {"role": "assistant", "content": reply},
        ]
    }

@app.post("/analyze")
async def analyze(
    file:    UploadFile = File(...),
    user_id: str        = Form("guest"),
    history: str        = Form("[]"),
):
    filename = file.filename or "upload"
    ext      = os.path.splitext(filename)[1].lower()
    try:
        prior_history = json.loads(history)
    except Exception:
        prior_history = []

    if ext in BLOCKED_EXTS:
        return {"analysis": "I am sorry, I cannot process audio or video files. Please upload your blood report as a PDF, Word document, Excel sheet, or image.", "filename": filename}
    if ext and ext not in SUPPORTED_EXTS:
        return {"analysis": f"I am not able to read .{ext.lstrip('.')} files right now. Please upload as PDF, DOCX, XLSX, CSV, TXT, or an image.", "filename": filename}

    try:
        file_bytes = await file.read()
        text       = extract_text(file_bytes, filename)

        if text == "[IMAGE_FILE]":
            analysis = analyse_image_with_groq(file_bytes, filename)
            save_chat(user_id, f"[Uploaded image: {filename}]", analysis)
            return {
                "analysis": analysis, "filename": filename,
                "new_history_entry": [
                    {"role": "user",      "content": f"[Uploaded blood report image: {filename}]"},
                    {"role": "assistant", "content": analysis},
                ]
            }

        if not text.strip():
            reply = call_groq("The user uploaded a file but no text could be extracted. Apologise kindly and ask them to try a PDF or DOCX or type the values in chat.", prior_history)
            return {"analysis": reply, "filename": filename}

        prompt = (
            f"The user uploaded a blood test report named '{filename}':\n\n{text[:5000]}\n\n"
            f"Analyse fully as a blood report specialist. Include all sections: "
            f"parameter analysis, overall summary, what to do next, food and nutrition recommendations, lifestyle tips. "
            f"Plain text only — no markdown, no ** bold, no ### headings, no tables."
        )
        analysis = call_groq(prompt, prior_history)
        save_chat(user_id, f"[Uploaded: {filename}]", analysis)
        return {
            "analysis": analysis, "filename": filename,
            "new_history_entry": [
                {"role": "user",      "content": f"[Uploaded blood report: {filename}]\n\n{text[:3000]}"},
                {"role": "assistant", "content": analysis},
            ]
        }
    except Exception as e:
        traceback.print_exc()
        return {"analysis": f"Something went wrong reading your file. Please try again. ({e})"}

class AnalyzeTextRequest(BaseModel):
    text:    str
    user_id: str  = "guest"
    history: list = []

@app.post("/analyze-text")
async def analyze_text(request: AnalyzeTextRequest):
    if not request.text.strip():
        raise HTTPException(400, "No text provided")
    prompt = (
        f"The user pasted blood report values:\n\n{request.text[:5000]}\n\n"
        f"Analyse fully. Include all sections. Plain text only — no markdown."
    )
    analysis = call_groq(prompt, request.history)
    save_chat(request.user_id, f"[Pasted report]\n\n{request.text[:2000]}", analysis)
    return {
        "analysis": analysis,
        "new_history_entry": [
            {"role": "user",      "content": f"[Pasted blood report values]\n\n{request.text[:3000]}"},
            {"role": "assistant", "content": analysis},
        ]
    }

@app.get("/history/{user_id}")
async def get_history(user_id: str):
    if not db or user_id == "guest":
        return {"messages": []}
    try:
        ref  = db.collection("chats").document(user_id).collection("messages")
        docs = ref.order_by("timestamp").stream()
        return {"messages": [{"role": d.get("role"), "content": d.get("content")} for d in docs]}
    except Exception as e:
        return {"messages": [], "error": str(e)}

@app.delete("/history/{user_id}")
async def clear_history(user_id: str):
    if not db or user_id == "guest":
        return {"cleared": False}
    try:
        ref = db.collection("chats").document(user_id).collection("messages")
        for d in ref.stream(): d.reference.delete()
        return {"cleared": True}
    except Exception as e:
        return {"cleared": False, "error": str(e)}

@app.post("/download-pdf")
async def download_pdf(request: PDFRequest):
    try:
        pdf_bytes = build_pdf(request.patient_name, request.conversation)
        buf = io.BytesIO(pdf_bytes)
        buf.seek(0)
        safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', request.patient_name)
        fn   = f"Vitalsy_Report_{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'}
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, f"PDF generation failed: {e}")

# ─────────────────────────────────────────────
# RUN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("brain:app", host="0.0.0.0", port=port, reload=False)